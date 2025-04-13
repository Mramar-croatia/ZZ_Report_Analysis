from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
import pandas as pd
from datetime import datetime

DATES_ROW = '7'

location_dict = {
    'P': 'Peščenica',
    'M': 'Špansko',
    'T': 'Trešnjevka',
    'J': 'Trešnjevka',
    'D': 'Dubrava',
    'Š': 'Špansko',
    'K': 'Kralj Tomislav'
}

DATES_ROW = '7'

def parse_date(date):
    """
    Parse a date from either a string or a datetime object and return a consistent format (DD.MM.YYYY).
    Handles multiple potential string formats and gracefully falls back for invalid dates.
    """
    if isinstance(date, str):
        # Try multiple common date formats
        date_formats = ["%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d", "%d/%m/%Y"]
        for fmt in date_formats:
            try:
                return datetime.strptime(date, fmt).strftime("%d.%m.%Y.")  # Standardize to DD.MM.YYYY
            except ValueError:
                continue
        date = date  # Return the original string if parsing fails for all formats
    
    elif isinstance(date, datetime):
        date =  date.strftime("%d.%m.%Y.")  # Convert datetime to DD.MM.YYYY
    
    if date[-4] == '.':
        date = date[:-2] + '02' + date[-2:]  # Remove trailing dot if present
    
    return date  # Return as-is if neither string nor datetime

def translate_school(school:  str) -> str:
    '''
    Translate school name to a standardized format
    
    Example:
    - 'St. John's' -> 'St.'
    - 'XV. gimnazija' -> 'XV.'
    - 'Gimnazija Kranj' -> 'GK'
    - 'Osnovna šola Ketteja in Murna' -> 'OSKIM'
    '''
    
    # If the school has a roman numeral, return the numeral
    if '.' in school:
        return school[0:school.index('.')+1]
    
    # Else, return the first letter of each word in the school name
    else:
        return ''.join([i[0].upper() for i in school.split(' ')])
    
def load_all_schools(sheet: object)  -> list[tuple[str, tuple[str]]]:
    '''
    Will load all schools from a passed sheet.
    
    Returns them in a list of tuples, where the first element is the school name and the second element is the range of the school name.
    
    Example:
    [('St.', ('A1', 'A2')), ('GK', ('A3', 'A5')), ('OSKIM', ('A6', 'A8'))]
    '''
    
    schools = []

    # Gets all the merged cell ranges, which contain the school names
    for merged_cell_range in sheet.merged_cells.ranges:
    
        # Get the range of the merged cell
        school_range = tuple(str(merged_cell_range).split(':'))
    
        # Translate the school name to a standardized format
        school_name = sheet[school_range[0]].value
        school_name = translate_school(school_name)
    
        # Append the school name and the range to the schools list in a tuple format
        schools.append((school_name, school_range))
    
    # Sort the schools by the row number of the first cell in the range (the left cell)
    schools = sorted(schools, key=lambda x: int(x[1][0][1:]))
        
    return schools

def load_classes(sheet: object, school_tuple: tuple[str, tuple[str]], next_school_start_row: int) -> list[tuple[str, str]]:
    '''
    Will load all classes from a passed sheet, for a passed school.
    Besides the sheet, the function requires a tuple containing the school name and the range of the school name. It also requires the row number of the next school so that it knows where to stop.
    Returns a list of tuples, where the first element is the class name and the second element is the start of the class range.
    
    Example:
    [('1. A', 'A2'), ('1. B', 'A3'), ('2. A', 'A4')]
    '''
    
    #  Get the school name and the start of the school range
    school_name, school_start = school_tuple[0], school_tuple[1][0]
    
    # Get the column and row of the cell where the school starts
    column, start_row = coordinate_from_string(school_start)
    start_row = int(start_row)
    
    classes = []
    
    # Iterate through the rows from the start of the school to the start of the next school
    for row in range(start_row+1, next_school_start_row):
        
        class_name = sheet[column+str(row)].value
        
        # If the cell is not empty, append the class name and the start of the class range to the classes list
        if class_name != None:
            classes.append((class_name, column+str(row)))
        
    return classes

def load_volunteers(sheet: object, location_name: str, school_name: str, class_tuple: tuple[str, str], next_class_start_row: int) -> pd.DataFrame:
    '''
    Will load all volunteers from a passed sheet, for a passed class (and school).
    
    Returns a DataFrame with the following columns:
    - volunteer_name
    - volunteer_dates
    - volunteer_class
    - volunteer_school
    '''
    
    # Initialize a dictionary to store the volunteers and to later turn it into a DataFrame
    volunteers_dict = {
        'volunteer_name': [],
        'volunteer_dates': [],
        'volunteer_class': [],
        'volunteer_school': []
    }
    
    # Get the class name and the start of the class range
    class_name, class_start = class_tuple
    
    # Get the column and row of the cell where the class starts
    column, start_row = coordinate_from_string(class_start)
    start_row = int(start_row)
    
    # Get the column letter of the first column, two columns to the right of the class column
    column = get_column_letter(column_index_from_string(column)+2)
    
    # Iterate through the rows from the start of the class to the start of the next class
    for row in range(start_row, next_class_start_row):
        
        # Get the volunteer name
        volunteer_name = sheet[column+str(row)].value
        
        # Get the dates where the volunteer had done some work
        matching_cells = [sheet[get_column_letter(cell.column)+DATES_ROW] for cell in sheet[row] if cell.value == 'da']
        dates = [parse_date(cell.value) for cell in matching_cells if cell.value is not None]
        
        hours = [(location_name, date) for date in dates]
                
        if volunteer_name != None:
            
            # Remove the * character from the volunteer name
            volunteer_name = volunteer_name.replace('*', '')
            
            volunteers_dict['volunteer_name'].append(volunteer_name)
            volunteers_dict['volunteer_dates'].append(hours)
            volunteers_dict['volunteer_class'].append(class_name)
            volunteers_dict['volunteer_school'].append(school_name)
        
    return pd.DataFrame(volunteers_dict)

def load_location(wb, location_sheet_name: str) -> pd.DataFrame:
    '''
    Will load all volunteers from a passed location sheet.
    
    Returns a DataFrame with the following columns:
    - volunteer_name
    - volunteer_dates
    - volunteer_class
    - volunteer_school
    '''
    
    print(f'Loading volunteers for {location_sheet_name}...')
    
    # Load the worksheet and the location name
    sheet = wb[location_sheet_name]
    location_name = location_dict[location_sheet_name]
    
    # Dictionary to store the volunteers for this location and turns it into a DataFrame
    this_location_dict = {
        'volunteer_name': [],
        'volunteer_dates': [],
        'volunteer_class': [],
        'volunteer_school': []
    }
    volunteer_df = pd.DataFrame(this_location_dict)
    
    # Load all schools and classes for this location
    schools = load_all_schools(sheet)
    
    # Iterate through all schools
    for school_tuple in schools:
        
        school_name, school_range = school_tuple
        
        # Get the row number of the next school, or set it to the last row of the sheet if it is the last school
        if schools.index((school_name, school_range)) == len(schools)-1:
            next_school_start_row = sheet.max_row+1
        else:
            next_school_start_row = int(coordinate_from_string(schools[schools.index((school_name, school_range))+1][1][0])[1])
        
        # Load all classes for this school
        classes = load_classes(sheet, school_tuple, next_school_start_row)
        
        # Iterate through all classes
        for class_tuple in classes:
            
            # Get the row number of the next class, or set it to the last row of the sheet if it is the last class
            if classes.index(class_tuple) == len(classes)-1:
                next_class_start_row = next_school_start_row
            else:
                next_class_start_row = int(classes[classes.index(class_tuple)+1][1][1:])
                
            # Load all volunteers for this class
            class_volunteer_df = load_volunteers(sheet, location_name, school_name, class_tuple, next_class_start_row)
            
            # Concatenate the class volunteers to the location volunteers
            volunteer_df = pd.concat([volunteer_df, class_volunteer_df], ignore_index=True)
            
    return volunteer_df

def load_year(file_path: str) -> object:
    '''
    Load the workbook, data_only=True to get the values of the formulas not the formulas themselves.
    '''
    
    # Load the workbook, data_only=True to get the values of the formulas not the formulas themselves
    wb = load_workbook(file_path, data_only=True)
    sheets = wb.sheetnames
    
    locations = []
    
    # Load all locations
    for sheet in sheets[1:]:
    
        location_df = load_location(wb, sheet)
        locations.append(location_df)
        
    
    year_df = pd.concat(locations, ignore_index=True)

    year_df = (
        year_df.groupby("volunteer_name", as_index=False)
        .agg({
            "volunteer_dates": lambda x: sum(x, []),  # Merge lists in 'volunteer_dates'
            **{col: "first" for col in year_df.columns if col not in ["volunteer_name", "volunteer_dates"]}  # Keep the first occurrence of other columns
        })
    )

    return year_df